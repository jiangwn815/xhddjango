from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict

from .models import Order, Task
from django.contrib.auth.models import User
from django.urls import reverse
import json, time
import requests
import os
import re
from bs4 import BeautifulSoup
from datetime import datetime
import calendar
import zipfile
import csv
import codecs, chardet
from openpyxl import Workbook, load_workbook
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login
from django.contrib.auth.hashers import make_password
import itchat


def index(request):
    '''
    latest_user_list = User.objects.order_by("-createdAt")
    print(latest_user_list[0].mobile)
    kk = {}
    for e in latest_user_list:
        kk[e.mobile] = e.createdAt
        print(kk)

    # template = loader.get_template('mainpages/index.html')
    context = {
        "latest_user_list": latest_user_list,
    }
    # return HttpResponse(template.render(context, request))
    '''
    # return render(request, 'mainpages/index.html', context)  # return a HttpResponse object
    return render(request, 'mainpages/index.html')


def register(request):
    return render(request, 'mainpages/register.html')


def login_page(request):
    return render(request, 'mainpages/login.html')


def login_view(request):
    username = request.POST['username']
    password = request.POST['password']
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        return HttpResponseRedirect(reverse('mainpages:index'))
    else:
        # Return an 'invalid login' error message.
        return render(request, 'mainpages/login.html')


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse('mainpages:index'))


def crawler(request):
    return render(request, 'mainpages/crawler.html')


# 下载文件（文件地址，文件名字，目标地址，文件大小阈值）
def downloadpic(picurl, picname, target_url, piclimit=100):
    headers = {"User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 10_3 like Mac OS X) AppleWebKit/602.1.50 (KHTML, like "
                             "Gecko) CriOS/56.0.2924.75 Mobile/14E5239e Safari/602.1",
               "Connection": "keep-alive",
               "Referer": target_url,
               "Pragma": "no-cache",
               "Host": "mtl.ttsqgs.com",
               "Accept-Encoding": "gzip, deflate, br",
               "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
               "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
               "Cache-Control": "no-cache"}
    size = 0

    try:
        picdata = requests.get(picurl, headers=headers, timeout=20)
        print("{0} len:{1:.2f} KB".format(picurl, len(picdata.content) / 1024))
    except requests.exceptions.ConnectionError:
        print(picname + " 无法下载")
        return size
    if len(picdata.content) / 1024 < piclimit:
        return size
    else:
        size = picdata.content/1024

    filename = picname
    while os.path.exists('./mainpages/download/' + filename):
        filename = filename.split(r'.')[0] + "-1." + filename.split(r'.')[1]
    with open('./mainpages/download/' + filename, 'wb') as picfile:
        picfile.write(picdata.content)
    return size


# 打开某个csv文件并写入传入的sheet
def mergexl(file,ws):
    with open(file) as csvfile:
        csvreader = csv.reader(csvfile)
        for x, row in enumerate(csvreader):
            for y, value in enumerate(row):
                ws.cell(row=x+1, column=y+1, value=value)


# 将某文件转换为UTF-8编码
def toutf8(fn):
    newfile = ""
    if os.path.isfile(fn) and fn.endswith('.csv') and "utf8" not in fn:
        zippath = '/'.join(fn.split(r'/')[0:-1])
        newfile = os.path.join(zippath, fn.split(r'.')[0] + "utf8.csv") # 创建带有utf8的文件名
        fncharset = chardet.detect(open(fn, "rb").read())  # 检测文件编码方式
        print("Open file:{0} Charset:{1}".format(fn, fncharset))
        print("New file:", newfile)
        with open(newfile, "wb") as csvutf8:  # w模式会把已存内容全部擦掉
            csvutf8.write(codecs.BOM_UTF8)  # 开头写入BOM防止win系统乱码

        with open(newfile, "a") as csvutf8:

            csvutf8writer = csv.writer(csvutf8, dialect='excel') # 准备读取csv文件
            with codecs.open(fn, "r", encoding="gbk") as csvfile: # 以gbk解码文件
                csvreader = csv.reader(csvfile)
                for row in csvreader:
                    csvutf8writer.writerow(row)
    if newfile:
        return newfile


# 解压某目录下所有压缩文件
def bjdata(request):
    ul = {"utf8file":{}}
    zippath = '/users/jwn/Desktop/工作文件/外呼/2017年3~11月下单妥投号码'

    for filename in os.listdir(zippath):  # 遍历目标目录下所有文件和文件夹
        fn = os.path.join(zippath, filename)
        utf8file = toutf8(fn)  # 转化当前文件
        if utf8file:
            ul["utf8file"][utf8file.split('.')[0].split(r'/')[-1]] = utf8file

    wb = Workbook()
    for filename, filepath in ul["utf8file"].items():
        ws = wb.create_sheet(filename) # 新建sheet
        mergexl(filepath, ws)
    wb.save(zippath+"/merge.xlsx")
    print("SHEET NAMES:", wb.sheetnames)
    return JsonResponse(ul)


def printstatus(res):
    print("Status code:", res.status_code)
    print("Encoding:", res.encoding)


def crawlerpic(request):
    ul = {}
    target_url = "https://m.meitulu.com/item/6932.html"
    home_r = requests.get(target_url)
    printstatus(home_r)

    print(request.GET["picnumber"])
    print(request.content_params)
    return JsonResponse(ul)
    home_soup = BeautifulSoup(home_r.text, 'lxml')
    url_list = home_soup.find('div', id='pages').find_all('a')[1:-1]

    for url in url_list:
        url = target_url.split(r'/item')[0]+url['href']
        print(url)

        r = requests.get(url)
        print("Headers", r.request.headers)
        img_soup = BeautifulSoup(r.text, 'lxml')
        img_lists = img_soup.find_all('img')
        # lists = soup.find_all('img', class_="content_img")

        for li in img_lists:
            picname = "".join(li['src'].split(r'/')[-2:])
            picurl = li['src']
            print(picurl)

            downloadpic(picurl, picname, target_url=target_url)

    return JsonResponse(ul)


def sms(request):
    user_list = User.objects
    order_list = Order.objects
    return render(request, 'mainpages/sms.html', {"ul": user_list, "ol": order_list})


def smslist(request):
    user_list = User.objects
    order_list = Order.objects
    return render(request, 'mainpages/sms_list.html', {"ul": user_list,"ol": order_list})


def smsedit(request):
    user_list = User.objects
    order_list = Order.objects
    return render(request, 'mainpages/sms_edit.html', {"ul": user_list,"ol": order_list})


def showuser(request, mobile):
    """
    try:
        user = User.objects.get(mobile=mobile)
    except User.DoesNotExist:
        raise Http404("User does not exist")
    """
    user = get_object_or_404(User, mobile=mobile)
    return render(request, 'mainpages/show_user.html', {"user": user})


def users(request):
    userslist = User.objects.order_by("date_joined")
    print(type(userslist))
    ul = {}
    for e in userslist:
        ul[e.mobile] = e.date_joined
    return JsonResponse(ul)


def createuser(request):
    User.objects.create(password=make_password(request.POST["password"]), username=request.POST["userName"])

    return HttpResponseRedirect(reverse('mainpages:index'))


def info(request, mobile):
    user = get_object_or_404(User, mobile=mobile)
    return render(request, 'mainpages/info.html', {'user': user})


def useramount(request, mobile):
    user = get_object_or_404(User, mobile=mobile)
    try:
        selected_order = user.order_set.get(out_trade_no=request.POST['order'])
    except(KeyError, Order.DoesNotExist):
        return render(request, 'mainpages/show_user.html', {
            'user': user,
            'error_message': "Order does not exist"
        })
    else:
        selected_order.total_fee = request.POST['amount']
        selected_order.save()
    return HttpResponseRedirect(reverse('mainpages:info', args=(user.mobile,)))


def tasks(request):
    task = Task.objects.order_by("-createdAt")
    ul = {}
    no = 1
    for e in task:
        ul[no] = {"id": str(e.id),
                  "createdAt": json.dumps(calendar.timegm(e.createdAt.timetuple())*1000),
                  "taskName": e.taskName,
                  "taskStatus": e.taskStatus}
        no = no + 1
    return JsonResponse(ul)


def createtask(request):

    sendtime = datetime.fromtimestamp(float(request.POST["sendTime"]))

    Task.objects.create(taskName=request.POST["task-name"],
                        taskContent=request.POST["task-content"],
                        startAt=sendtime,
                        taskStatus=request.POST["taskStatus"],
                        senderNumber=request.POST["sender-number"],
                        receiverNumber=request.POST["receiver-number"]
                        )
    return HttpResponseRedirect(reverse('mainpages:smslist'))


def viewtask(request):
    print(request.GET['id'])
    task = get_object_or_404(Task, id=request.GET['id'])
    print(task.taskName)
    print(task.createdAt)
    return JsonResponse({
        "id": task.id,
        "taskName": task.taskName,
        "taskContent": task.taskContent,
        "taskStatus": task.taskStatus,
        "startAt": json.dumps(calendar.timegm(task.startAt.timetuple())*1000),
        "createdAt": json.dumps(calendar.timegm(task.createdAt.timetuple())*1000),
        "senderNumber": task.senderNumber,
        "receiverNumber": task.receiverNumber
    })


@csrf_exempt
def deletetask(request):
    task = get_object_or_404(Task, id=request.POST['id'])
    task.delete()
    return JsonResponse({"code": "0"})


