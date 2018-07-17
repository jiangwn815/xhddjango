from django.shortcuts import get_object_or_404, render
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template import loader
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db.models import Avg
from django.db.models import Count
import os, csv
import copy
import csv,chardet,codecs
import xlrd
from .models import Productinfo,TeleUser, Bill, ResourceUsage, KDUser
from fileprocess.models import UploadFile
from openpyxl import Workbook, load_workbook
from datetime import datetime, date, timedelta
import time as ttt
from django.forms.models import model_to_dict
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from django.conf import settings
import itchat
from .mapping import tele_field_mapping
import re
from .models import TeleDepartment

def index(request):

    files = UploadFile.objects.filter(user=request.user)
    for file in files.all():
        print(file.filedata.__str__())
        print(file.filedata.path)
        print(file.filedata.url)
        print(file.filedata.name)
        print(file.id)
    return render(request, 'datacleaning/index.html', {"files":files})


def userlist(request):
    users = Productinfo.objects.all()[:100]
    template = loader.get_template("datacleaning/user.html")
    context = {
        "users": users
    }
    return HttpResponse(template.render(context, request))


def open_QR():
    for get_count in range(10):
        print('Getting uuid')
        uuid = itchat.get_QRuuid()
        while uuid is None:
            uuid = itchat.get_QRuuid();
            time.sleep(1)
        print('Getting QR Code')
        if itchat.get_QR(uuid):
            break
        elif get_count >= 9:
            print('Failed to get QR Code, please restart the program')

    print('Please scan the QR Code')
    return uuid


def wechat(request):
    return render(request, 'datacleaning/wechat.html')


def getqrimg(request):
    ul = {}
    print('Getting oepnqr')
    uuid = open_QR()
    waitForConfirm = False
    while 1:
        status = itchat.check_login(uuid)
        if status == '200':
            break
        elif status == '201':
            if waitForConfirm:
                print('Please press confirm')
                waitForConfirm = True
        elif status == '408':
            print('Reloading QR Code')
            uuid = open_QR()
            waitForConfirm = False
        time.sleep(3)
    userInfo = itchat.web_init()
    print("ui",userInfo)
    # itchat.show_mobile_login()
    # itchat.get_contract()
    # print('Login successfully as %s' % userInfo['NickName'])
    itchat.start_receiving()
    return JsonResponse(ul)
    # itchat.send('Hello, filehelper', toUserName='filehelper')


def userlist_paginator(request):
    st = request.GET.get('searchText', default="")
    c6 = request.GET.get('searchChannelSix', default="")
    user_list = TeleUser.objects.all()
    if st:
        user_list = user_list.filter(charge_plan__contains=st)
    if c6:
        user_list = user_list.filter(seller_channel_sixth__contains=c6)

    sumno = user_list.count()
    paginator = Paginator(user_list, 15)
    page = request.GET.get('page')
    users = paginator.get_page(page)
    return render(request, 'datacleaning/userpaginator.html', {'contacts': users,
                                                               'sumno': sumno})


@login_required()
@permission_required('datamining.access_teledata', raise_exception=True)
def showname(request):
    st = request.GET.get('searchText', default="")
    c6 = request.GET.get('searchChannelSix', default="")
    sn = request.GET.get('searchName', default="")
    user_list = TeleUser.objects.order_by('username').values('username').distinct()
    if st:
        user_list = user_list.filter(charge_plan__contains=st)
    if c6:
        user_list = user_list.filter(seller_channel_sixth__contains=c6)
    if sn:
        user_list = user_list.filter(username__contains=sn)
    sumno = user_list.count()
    paginator = Paginator(user_list, 15)
    page = request.GET.get('pagelist')
    users = paginator.get_page(page)
    return render(request, 'datacleaning/name_list.html', {'contacts': users, 'sumno': sumno})


@login_required()
@permission_required('datamining.access_teledata', raise_exception=True)
def product_info_list(request):
    username = request.user.__str__()
    print(username)
    st = request.GET.get('searchText', default="")
    c6 = request.GET.get('searchChannelSix', default="")
    sn = request.GET.get('searchName', default="")
    pt = request.GET.get('searchProductType', default="")
    sp = request.GET.get('searchSubscribePlan', default="")
    cp = request.GET.get('searchChargePlan', default="")
    productlist = Productinfo.objects.order_by('user_no').all()
    if st:
        productlist = productlist.filter(charge_plan__contains=st)
    if c6:
        productlist = productlist.filter(seller_channel_sixth__contains=c6)
    if sn:
        productlist = productlist.filter(username__contains=sn)
    if pt:
        if pt == 'h':
            productlist = productlist.filter(mainproduct_second__contains="后付费语音")
        elif pt == 'y':
            productlist = productlist.filter(mainproduct_second__contains="预付费语音")
    subscribe_plan = list(productlist.order_by('subscribe_plan').values('subscribe_plan').distinct())
    subscribe_plan_list = [x['subscribe_plan'].split('(')[0] for x in subscribe_plan]
    if sp:
        productlist = productlist.filter(subscribe_plan__contains=sp)
    charge_plan = list(productlist.order_by('charge_plan').values('charge_plan').distinct())
    charge_plan_list = [x['charge_plan'].split('(')[0] for x in charge_plan]
    print(subscribe_plan_list)
    product_no = productlist.count()
    paginator = Paginator(productlist, 15)
    page = request.GET.get('pagelist')
    users = paginator.get_page(page)
    return render(request, 'datacleaning/product_info_list.html', {'productlist': users,
                                                                   'subscribe_plan': subscribe_plan_list,
                                                                   'charge_plan': charge_plan_list,
                                                                   'product_no': product_no,
                                                                   'current_user': username
                                                                   })


def showcustomer(request, customer_id):
    template = loader.get_template("datacleaning/customer_info.html")
    customer_name_list = []
    user_no_list = []
    for customer in TeleUser.objects.filter(customer_id=customer_id):
        if customer.username:
            customer_name_list.append(customer.username)
            if customer.user_no not in user_no_list:
                user_no_list.append(customer.user_no)
    print(customer_name_list)
    print(user_no_list)
    user_count = TeleUser.objects.filter(customer_id=customer_id).annotate(no_user=Count('user_no'))
    print(user_count[0].no_user)
    zz_income = Bill.objects.filter(customer_id=customer_id).aggregate(Avg('zz_income'))
    ru = ResourceUsage.objects.filter(customer_id=customer_id).last()
    context = {
        "customer_names": set(customer_name_list),
        "user_nos": set(user_no_list),
        "bill": zz_income,
        "ru": ru
    }
    print("Context", context)
    return HttpResponse(template.render(context, request))


def showuser(request, user_no):
    template = loader.get_template("datacleaning/user_info.html")
    user_count = TeleUser.objects.filter(user_no=user_no).annotate(no_user=Count('user_no'))
    print(user_count[0].no_user)
    zz_avg = Bill.objects.filter(user_no=user_no).aggregate(Avg('zz_income'))
    ru = ResourceUsage.objects.filter(user_no=user_no).last()
    user = TeleUser.objects.get(user_no=user_no)
    context = {
        "zz_avg": zz_avg,
        "ru": ru,
        'user': user,
        'zz_avg': zz_avg['zz_income__avg']
    }
    print("Context", context)
    return HttpResponse(template.render(context, request))


# 打开某个csv文件并写入传入的sheet
def mergexl(file, ws):
    with open(file) as csvfile:
        csvreader = csv.reader(csvfile)
        for x, row in enumerate(csvreader):
            for y, value in enumerate(row):
                ws.cell(row=x+1, column=y+1, value=value)


# 将某文件转换为UTF-8编码
def toutf8(fn):
    newfile = ""
    if os.path.isfile(fn) and fn.endswith('.csv') and "utf8" not in fn:
        zippath = '/'.join(fn.split(r'/')[0:-1])
        newfile = os.path.join(zippath, fn.split(r'.')[0] + "utf8.csv") # 创建带有utf8的文件名
        fncharset = chardet.detect(open(fn, "rb").read())  # 检测文件编码方式
        print("Open file:{0} Charset:{1}".format(fn, fncharset))
        if fncharset["encoding"] not in ("GB2312", "gbk"):
            return
        with open(newfile, "wb") as csvutf8:  # w模式会把已存内容全部擦掉
            csvutf8.write(codecs.BOM_UTF8)  # 开头写入BOM防止win系统乱码

        with open(newfile, "a") as csvutf8:

            csvutf8writer = csv.writer(csvutf8, dialect='excel') # 准备读取csv文件
            with codecs.open(fn, "r", encoding="gbk") as csvfile: # 以gbk解码文件
                print("WORKING ON:", fn)
                csvreader = csv.reader(csvfile)
                for row in csvreader:
                    print(row)
                    csvutf8writer.writerow(row)
    if newfile:
        return newfile


# 返回表中第一行各字段的列值
# 参数1：表格第一行的表头
def field_location_mapping(first_row):
    location = {}
    for cell in first_row:  # 提取表的第一行
        print("开始遍历文件头......", cell.value)
        for key, field in tele_field_mapping.items():  # 遍历对应字典中各项
            if isinstance(field, str):
                if cell.value.endswith(field):
                    location[key] = cell.col_idx  # 存储某个字段在表中坐标
            elif isinstance(field, dict):      # 如果表中字段为多时间维度
                for k, v in field.items():
                    if cell.value.endswith(v):
                        if key in location:
                            location[key].update({k: cell.col_idx})
                        else:
                            location.update({key: {k: cell.col_idx}})
    return location


def dealxlsx(request):
    ul = {}
    file_path = UploadFile.objects.get(pk=request.GET.get('fileSelect')).filedata.path  # 获取用户所选文件的地址
    print('开始处理文件：', file_path)
    start = ttt.time()
    with open(file_path, 'r') as cf:
        reader = csv.reader(cf)
        count = 0
        for row in reader:
            KDUser.objects.update_or_create(user_no=row[0], mobile_no=row[1])
            count = count+1
            if count == 10000:
                print(count, " Finished")
                break
    end = ttt.time()
    print(end-start)
    return JsonResponse(ul)


    ws = load_workbook(file_path).active  # 获取文件中的活动sheet
    mapping = field_location_mapping(ws[1])
    rc = 0
    for row in ws.iter_rows(min_row=2, max_row=500):
        ui = {}  # 存储单行字段-值信息
        rc = rc+1
        if rc%10000 ==0:
            print('已处理', rc, " 行数据")
        for cell in row:
            for key, idx in mapping.items():
                if isinstance(idx, int):
                    if cell.col_idx == idx:
                        ui[key] = cell.value
                elif isinstance(idx, dict):
                    for k, v in idx.items():
                        if cell.col_idx == v:
                            if key in ui:
                                ui[key].update({k: cell.value})
                            else:
                                ui.update({key: {k: cell.value}})

        KDUser.objects.update_or_create(user_no=ui['user_no'], mobile_no=ui['mobile_no'], defaults={'type': 'kuandai20180717'})
        '''
        try:
            sc2 = re.match(r'^(\w+?)\((\d+)\)', ui['seller_channel_second'])
            sc3 = re.match(r'^(\w+?)\((\d+)\)', ui['seller_channel_third'])
            sc4 = re.match(r'^(\w+?)\((\d+)\)', ui['seller_channel_fourth'])
            sc5 = re.match(r'^(\w+?)\((\d+)\)', ui['seller_channel_fifth'])
        except TypeError as te:
            print("Error:", te)
            print(ui['seller_channel_second'], ui['seller_channel_third'], ui['seller_channel_fourth'])
        if sc2 and sc3 and sc4 and sc5:
            o2, c2 = TeleDepartment.objects.update_or_create(name=sc2.group(1), department_id=sc3.group(2), level=2)
            o3, c3 = TeleDepartment.objects.update_or_create(name=sc3.group(1), department_id=sc3.group(2), level=3,
                                                             defaults={'superior': o2})
            o4, c4 = TeleDepartment.objects.update_or_create(name=sc4.group(1), department_id=sc4.group(2), level=4,
                                                             defaults={'superior': o3})
            o5, c5 = TeleDepartment.objects.update_or_create(name=sc5.group(1), department_id=sc5.group(2), level=5,
                                                             defaults={'superior': o4})
        '''
    print("{} obejcts created".format(rc))
    return JsonResponse(ul)


def dealxlsx2(request):
    print("开始处理文件......")
    ul = {}

    field_mapping = tele_field_mapping
    fpath = ""
    start = ttt.time()
    wb = load_workbook(fpath)
    stop = ttt.time()
    print("文件载入时间：", stop-start)
    ul['file_load_time'] = stop-start
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 利用表头确定每列数据对应的字段
        multi_rows = False

        users = field_location_mapping(ws[1], field_mapping)

        print("文件头匹配时间：", stop - start)
        ul['file_match_time'] = stop-start

        for row in ws.iter_rows(min_row=2):
            ui = {}  # 存储单行字段-值信息
            for cell in row:
                for key, idx in users.items():
                    if isinstance(idx, int):
                        if cell.col_idx == idx:
                            ui[key] = cell.value
                    elif isinstance(idx, dict):
                        for k, v in idx.items():
                            if cell.col_idx == v:
                                if key in ui:
                                    ui[key].update({k: cell.value})
                                else:
                                    ui.update({key: {k: cell.value}})

            if multi_rows:
                data_row = {}
                ITEM_NUMBER = 6
                current_month = 2
                for i in range(0, ITEM_NUMBER):
                    if current_month-i > 0:
                        account_date = date(2018, current_month-i, 1)
                    else:
                        account_date = date(2017, current_month-i+12, 1)
                    for key, field in ui.items():
                        if isinstance(field, dict):
                            data_row[key] = field[key+"_last"+str(i)]
                        elif isinstance(field, (str, int)):
                            data_row[key] = field
                    data_row['account_date'] = account_date
                    bill_db = copy.deepcopy(data_row)
                    ru_db = copy.deepcopy(data_row)
                    bill_db.pop('call_times')
                    bill_db.pop('data_usage')
                    bill_db.pop('text_usage')
                    ru_db.pop('zz_income')
                    print('bill value:', bill_db)
                    Productinfo.objects.update_or_create(
                        defaults=bill_db,
                        user_no=bill_db['user_no'], account_date=bill_db['account_date'])
                    #ResourceUsage.objects.update_or_create(
                       # defaults=ru_db,
                        #user_no=bill_db['user_no'], account_date=bill_db['account_date'])

            else:
                Productinfo.objects.update_or_create(defaults={'user_no': ui['user_no']}, **ui)
        stop = time()
        print("数据处理时间：", stop - start)
        ul['file_process_time'] = stop-start
    return JsonResponse(ul)


# 汇总某目录下所有压缩文件
def bjdata(request):
    ul = {"utf8file": {}}
    zippath = '/users/jwn/Desktop/工作文件/外呼/2017年3~11月下单妥投号码'

    for filename in os.listdir(zippath):  # 遍历目标目录下所有文件和文件夹
        fn = os.path.join(zippath, filename)
        utf8file = toutf8(fn)  # 转化当前文件到utf8格式
        if utf8file:
            ul["utf8file"][utf8file.split('.')[0].split(r'/')[-1]] = utf8file

    wb = Workbook()
    for filename, filepath in ul["utf8file"].items():
        ws = wb.create_sheet(filename)  # 新建sheet
        mergexl(filepath, ws)

    wb.save(zippath+"/merge.xlsx")
    print("SHEET NAMES:", wb.sheetnames)

    return JsonResponse(ul)


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active()
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


def collectxls(request):
    ul = {}
    path = '/Users/jwn/Desktop/工作文件/外呼/佣金建档/'
    xls = [file for file in os.listdir(path) if file.endswith('.xls')]
    wb = Workbook()
    for filename in xls:  # 遍历目标目录下所有文件和文件夹
        fn = os.path.join(path, filename)
        ws = wb.create_sheet(filename.split('.')[0])
        print(filename.split('.')[0])
        ws_copy = xlrd.open_workbook(fn).sheet_by_index(0)
        for row in range(0, ws_copy.nrows):
            for col in range(0, ws_copy.ncols):
                ws.cell(row=row+1, column=col+1, value=ws_copy.cell_value(row, col))
    wb.save(path+'output.xls')
    ul["msg"] = "success!"
    return JsonResponse(ul)
